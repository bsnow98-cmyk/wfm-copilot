"""
Excel report export.

Builds a multi-sheet workbook for one forecast run, including any staffing
scenarios that have been computed against it. The workbook has charts and is
formatted for review in Excel/Numbers/Google Sheets — the data ops user's
native tools.

Sheets produced:
    Summary           — run metadata, MAPE/WAPE, sheet index
    Forecast          — per-interval offered + AHT, with a line chart
    Staff_SLxx_ATyy_Shzz   — one sheet per staffing scenario, with chart
                              (sheet name encodes the parameters)

We use openpyxl directly (not pandas to_excel) so we can drop charts in.
"""
from __future__ import annotations

from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session

# --- styling ----------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
LABEL_FONT = Font(bold=True)
DATETIME_FMT = "yyyy-mm-dd hh:mm"
PCT_FMT = "0.00%"


def _style_header_row(ws, headers: Sequence[str]) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def _autosize(ws, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# --- the public entry point -------------------------------------------
def build_forecast_report(db: Session, forecast_run_id: int) -> bytes:
    """Returns the bytes of an .xlsx workbook for the given forecast run."""

    fc = db.execute(
        text("""
            SELECT id, queue, channel, model_name, status,
                   horizon_start, horizon_end, mape, wape,
                   created_at, started_at, completed_at, error_message
            FROM forecast_runs WHERE id = :id
        """),
        {"id": forecast_run_id},
    ).mappings().first()
    if fc is None:
        raise ValueError(f"forecast_run_id {forecast_run_id} not found")

    intervals = db.execute(
        text("""
            SELECT interval_start, forecast_offered, forecast_aht_seconds
            FROM forecast_intervals
            WHERE forecast_run_id = :id
            ORDER BY interval_start
        """),
        {"id": forecast_run_id},
    ).mappings().all()

    staffings = db.execute(
        text("""
            SELECT id, service_level_target, target_answer_seconds,
                   target_asa_seconds, shrinkage, created_at
            FROM staffing_requirements
            WHERE forecast_run_id = :id
            ORDER BY created_at
        """),
        {"id": forecast_run_id},
    ).mappings().all()

    wb = Workbook()
    _build_summary_sheet(wb, fc, intervals, staffings)
    _build_forecast_sheet(wb, intervals)
    for s in staffings:
        rows = db.execute(
            text("""
                SELECT interval_start, forecast_offered, forecast_aht_seconds,
                       required_agents_raw, required_agents,
                       expected_service_level, expected_asa_seconds, occupancy
                FROM staffing_requirement_intervals
                WHERE staffing_id = :id
                ORDER BY interval_start
            """),
            {"id": s["id"]},
        ).mappings().all()
        _build_staffing_sheet(wb, s, rows)

        # If any schedules were solved against this staffing, add a coverage sheet.
        schedules = db.execute(
            text("""
                SELECT id, name, solver_status, objective_value,
                       total_understaffed_intervals
                FROM schedules
                WHERE staffing_id = :id
                ORDER BY created_at
            """),
            {"id": s["id"]},
        ).mappings().all()
        for sch in schedules:
            cov_rows = db.execute(
                text("""
                    SELECT interval_start, required_agents,
                           scheduled_agents, shortage
                    FROM schedule_coverage
                    WHERE schedule_id = :id
                    ORDER BY interval_start
                """),
                {"id": sch["id"]},
            ).mappings().all()
            if cov_rows:
                _build_coverage_sheet(wb, sch, cov_rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- sheet builders ---------------------------------------------------
def _build_summary_sheet(wb: Workbook, fc, intervals, staffings) -> None:
    ws = wb.active
    ws.title = "Summary"

    # Title bar
    ws["A1"] = "WFM Copilot — Forecast Report"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 30

    rows = [
        ("Forecast ID", fc["id"]),
        ("Queue", fc["queue"]),
        ("Channel", fc["channel"]),
        ("Model", fc["model_name"]),
        ("Status", fc["status"]),
        ("Horizon start", _fmt(fc["horizon_start"])),
        ("Horizon end", _fmt(fc["horizon_end"])),
        ("MAPE (lower is better)", float(fc["mape"]) if fc["mape"] is not None else None),
        ("WAPE (lower is better)", float(fc["wape"]) if fc["wape"] is not None else None),
        ("Forecast intervals", len(intervals)),
        ("Staffing scenarios", len(staffings)),
        ("Created at", _fmt(fc["created_at"])),
        ("Completed at", _fmt(fc["completed_at"])),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=i, column=2, value=value)
        if isinstance(value, float):
            c.number_format = PCT_FMT

    # Sheet index
    next_row = 3 + len(rows) + 2
    ws.cell(row=next_row, column=1, value="Sheets in this workbook").font = LABEL_FONT
    next_row += 1
    ws.cell(row=next_row, column=1, value="Forecast")
    ws.cell(row=next_row, column=2, value="per-interval volume + AHT, with chart")
    for s in staffings:
        next_row += 1
        sheet_name = _staffing_sheet_name(s)
        params = _staffing_params_label(s)
        ws.cell(row=next_row, column=1, value=sheet_name)
        ws.cell(row=next_row, column=2, value=params)

    _autosize(ws, {"A": 28, "B": 36})


def _build_forecast_sheet(wb: Workbook, intervals) -> None:
    ws = wb.create_sheet("Forecast")
    headers = ["interval_start", "forecast_offered", "forecast_aht_seconds"]
    _style_header_row(ws, headers)

    for i, iv in enumerate(intervals, start=2):
        ws.cell(row=i, column=1, value=_naive(iv["interval_start"])).number_format = DATETIME_FMT
        ws.cell(row=i, column=2, value=float(iv["forecast_offered"]))
        ws.cell(row=i, column=3, value=float(iv["forecast_aht_seconds"]))

    _autosize(ws, {"A": 18, "B": 18, "C": 22})
    ws.freeze_panes = "A2"

    if intervals:
        chart = LineChart()
        chart.title = "Forecast offered volume"
        chart.x_axis.title = "Interval"
        chart.y_axis.title = "Offered"
        chart.height = 10
        chart.width = 28
        data = Reference(ws, min_col=2, min_row=1,
                         max_row=len(intervals) + 1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2,
                         max_row=len(intervals) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")


def _build_staffing_sheet(wb: Workbook, s, rows) -> None:
    sheet_name = _staffing_sheet_name(s)
    ws = wb.create_sheet(sheet_name)

    headers = [
        "interval_start", "forecast_offered", "forecast_aht_seconds",
        "required_agents_raw", "required_agents",
        "expected_service_level", "expected_asa_seconds", "occupancy",
    ]
    _style_header_row(ws, headers)

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=_naive(r["interval_start"])).number_format = DATETIME_FMT
        ws.cell(row=i, column=2, value=float(r["forecast_offered"]))
        ws.cell(row=i, column=3, value=float(r["forecast_aht_seconds"]))
        ws.cell(row=i, column=4, value=int(r["required_agents_raw"]))
        ws.cell(row=i, column=5, value=int(r["required_agents"]))
        sl_cell = ws.cell(row=i, column=6,
                          value=float(r["expected_service_level"]) if r["expected_service_level"] is not None else None)
        sl_cell.number_format = PCT_FMT
        ws.cell(row=i, column=7,
                value=float(r["expected_asa_seconds"]) if r["expected_asa_seconds"] is not None else None)
        occ_cell = ws.cell(row=i, column=8,
                           value=float(r["occupancy"]) if r["occupancy"] is not None else None)
        occ_cell.number_format = PCT_FMT

    _autosize(ws, {"A": 18, "B": 16, "C": 20, "D": 20, "E": 18,
                   "F": 22, "G": 20, "H": 14})
    ws.freeze_panes = "A2"

    if rows:
        chart = LineChart()
        chart.title = f"Required agents — {_staffing_params_label(s)}"
        chart.x_axis.title = "Interval"
        chart.y_axis.title = "Agents"
        chart.height = 10
        chart.width = 28
        data = Reference(ws, min_col=5, min_row=1,
                         max_row=len(rows) + 1, max_col=5)
        cats = Reference(ws, min_col=1, min_row=2,
                         max_row=len(rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "J2")


def _build_coverage_sheet(wb: Workbook, sch, cov_rows) -> None:
    """Schedule coverage: required vs. scheduled vs. shortage per interval,
    with a chart overlaying the two curves so you can see gaps at a glance."""
    sheet_name = f"Cov_sched_{int(sch['id'])}"[:31]
    ws = wb.create_sheet(sheet_name)

    headers = ["interval_start", "required_agents", "scheduled_agents", "shortage"]
    _style_header_row(ws, headers)

    for i, r in enumerate(cov_rows, start=2):
        ws.cell(row=i, column=1, value=_naive(r["interval_start"])).number_format = DATETIME_FMT
        ws.cell(row=i, column=2, value=int(r["required_agents"]))
        ws.cell(row=i, column=3, value=int(r["scheduled_agents"]))
        ws.cell(row=i, column=4, value=int(r["shortage"]))

    _autosize(ws, {"A": 18, "B": 18, "C": 18, "D": 14})
    ws.freeze_panes = "A2"

    # Add a brief metadata block to the right of the data.
    meta_col = 6
    ws.cell(row=1, column=meta_col, value="Schedule").font = LABEL_FONT
    ws.cell(row=1, column=meta_col + 1, value=sch["name"])
    ws.cell(row=2, column=meta_col, value="Solver status").font = LABEL_FONT
    ws.cell(row=2, column=meta_col + 1, value=sch["solver_status"])
    ws.cell(row=3, column=meta_col, value="Objective").font = LABEL_FONT
    ws.cell(row=3, column=meta_col + 1, value=float(sch["objective_value"]) if sch["objective_value"] is not None else None)
    ws.cell(row=4, column=meta_col, value="Understaffed intervals").font = LABEL_FONT
    ws.cell(row=4, column=meta_col + 1, value=int(sch["total_understaffed_intervals"]) if sch["total_understaffed_intervals"] is not None else None)
    ws.column_dimensions[get_column_letter(meta_col)].width = 24
    ws.column_dimensions[get_column_letter(meta_col + 1)].width = 28

    if cov_rows:
        chart = LineChart()
        chart.title = f"Required vs. Scheduled — {sch['name']}"
        chart.x_axis.title = "Interval"
        chart.y_axis.title = "Agents"
        chart.height = 10
        chart.width = 28
        # Two series: required (col 2) and scheduled (col 3)
        data = Reference(ws, min_col=2, min_row=1,
                         max_row=len(cov_rows) + 1, max_col=3)
        cats = Reference(ws, min_col=1, min_row=2,
                         max_row=len(cov_rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "F8")


# --- helpers ----------------------------------------------------------
def _fmt(dt) -> str | None:
    return dt.strftime("%Y-%m-%d %H:%M") if dt is not None else None


def _naive(dt):
    """openpyxl rejects timezone-aware datetimes ("Excel does not support
    timezones..."). Postgres TIMESTAMPTZ comes back as tz-aware UTC, so we
    drop the tzinfo before writing into a cell. We treat the wall-clock UTC
    time as the value, which matches how every WFM tool displays intervals."""
    if dt is None:
        return None
    return dt.replace(tzinfo=None) if dt.tzinfo is not None else dt


def _staffing_sheet_name(s) -> str:
    """Encode parameters into a 31-char-max sheet name. Includes ASA when set."""
    parts = ["Staff"]
    if s.get("service_level_target") is not None:
        parts.append(f"SL{int(float(s['service_level_target']) * 100)}")
        parts.append(f"AT{int(s['target_answer_seconds'])}")
    if s.get("target_asa_seconds") is not None:
        parts.append(f"ASA{int(s['target_asa_seconds'])}")
    parts.append(f"Sh{int(float(s['shrinkage']) * 100)}")
    return "_".join(parts)[:31]


def _staffing_params_label(s) -> str:
    """Human-readable summary like 'SL≥80% in 20s, ASA<=30s, shrinkage 30%'."""
    bits: list[str] = []
    if s.get("service_level_target") is not None:
        bits.append(
            f"SL≥{float(s['service_level_target']):.0%} in "
            f"{int(s['target_answer_seconds'])}s"
        )
    if s.get("target_asa_seconds") is not None:
        bits.append(f"ASA≤{int(s['target_asa_seconds'])}s")
    bits.append(f"shrinkage {float(s['shrinkage']):.0%}")
    return ", ".join(bits)
